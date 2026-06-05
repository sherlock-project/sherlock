"""
Excel Raporlama Modulu
openpyxl ile Excel export
"""

from pathlib import Path
from typing import List
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from sherlock_project.result import QueryResult


class ExcelReporter:
    """Excel rapor uretici"""

    def __init__(self):
        self.header_fill = PatternFill(start_color='1f538d', end_color='1f538d', fill_type='solid')
        self.header_font = Font(color='FFFFFF', bold=True, size=12)

    def generate(
        self,
        username: str,
        results: List[QueryResult],
        output_path: str
    ) -> str:
        """
        Excel raporu olustur

        Args:
            username: Aranan kullanici adi
            results: Tarama sonuclari
            output_path: Cikis dosya yolu

        Returns:
            Olusturulan dosya yolu
        """
        wb = Workbook()
        ws = wb.active
        ws.title = f'{username[:20]} Results'

        # Baslik
        ws['A1'] = 'Sherlock Report'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f'Username: {username}'
        ws['A3'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'

        # Tablo basliklari
        headers = ['Site Name', 'Status', 'URL', 'Response Time (s)', 'HTTP Status']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')

        # Veriler
        row = 6
        for result in results:
            ws.cell(row=row, column=1, value=result.site_name)
            ws.cell(row=row, column=2, value=result.status.value)
            ws.cell(row=row, column=3, value=result.site_url_user)
            ws.cell(row=row, column=4, value=round(result.query_time, 3))
            ws.cell(row=row, column=5, value=result.context.get('http_status', ''))
            row += 1

        # Kolon genislikleri
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15

        # Ozet
        ws[f'A{row + 2}'] = 'Summary'
        ws[f'A{row + 2}'].font = Font(bold=True)
        found_count = sum(1 for r in results if r.status.value == 'Claimed')
        ws[f'A{row + 3}'] = f'Found: {found_count}'
        ws[f'A{row + 4}'] = f'Total Checked: {len(results)}'

        wb.save(output_path)
        return output_path
